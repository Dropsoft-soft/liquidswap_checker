import os
import json, time
from loguru import logger
import requests
from openpyxl import Workbook
from datetime import datetime
import openpyxl
'''Settings use or not use proxies and filename'''
EXEL_FILENAME = 'liquidswap.xlsx'
USE_PROXY = True

with open("addresses.txt", "r") as f:
    WALLETS = [row.strip() for row in f]

with open("proxies.txt", "r") as f:
    PROXIES = [row.strip() for row in f]

def get_wallet_proxies(wallets, proxies):
    try:
        result = {}
        for i in range(len(wallets)):
            result[wallets[i]] = proxies[i % len(proxies)]
        return result
    except: None

WALLET_PROXIES  = get_wallet_proxies(WALLETS, PROXIES)


def request(method="get", request_retry=0, wallet=0, proxy=None, **kwargs):
    session = requests.Session()

    if proxy is not None:
        session.proxies.update(
            {
                "http": f"{proxy}",
                "https": f"{proxy}"
            }
        )

    if request_retry > 4:
        return
    retry = 0
    while True:
        try:
            if method == "post":
                response = session.post(**kwargs, verify=False)
            elif method == "get":
                response = session.get(**kwargs, verify=False)
            elif method == "put":
                response = session.put(**kwargs, verify=False)
            elif method == "options":
                response = session.options(**kwargs, verify=False)

            logger.info(f'{wallet}, status_code {response.status_code} response: {response}')

            if response.status_code == 201 or response.status_code == 200:
                time.sleep(5)
                try:
                    return response.json()
                except json.decoder.JSONDecodeError:
                    logger.info('The request success but not contain a JSON')
                    break
            else:
                logger.error(f'[{wallet} - Bad status code: {response.status_code} {response.json()}')
                time.sleep(15)
                retry += 1
                if retry > 4:
                    break

        except Exception as error:
            logger.error(f'{wallet} - {kwargs["url"]} failed to make request | {error}')
            time.sleep(15)
            request(method=method, request_retry=request_retry + 1, wallet=wallet, proxy=proxy, **kwargs)
            break

def set_column_widths(sheet):
    column_widths = {
        'A': 50,  # Wallet
        'B': 15,  # 
        'C': 15,  # Amount
        'D': 15,  # Points
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

def add_data(wallet: str, Amount1: str, Amount2: str, Amount3: str):
    file_name = EXEL_FILENAME
    if not os.path.exists(file_name):
        book = Workbook()
        sheet = book.active
        sheet['A1'] = 'Wallet'
        sheet['B1'] = 'Amount1'
        sheet['C1'] = 'Amount2'
        sheet['D1'] = 'Amount3'
        set_column_widths(sheet)
        book.save(file_name)
        book.close()

    book = openpyxl.load_workbook(file_name)
    sheet = book.active

    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1).value = wallet
    sheet.cell(row=new_row, column=2).value = Amount1
    sheet.cell(row=new_row, column=3).value = Amount2
    sheet.cell(row=new_row, column=4).value = Amount3

    book.save(file_name)
    book.close()
    logger.success(f'Data added to {file_name}: {wallet}, {Amount1}, {Amount2}, {Amount3}')

def start_check(wallet):
    proxy = None
    if USE_PROXY:
        proxy = WALLET_PROXIES[wallet]

    url = f'https://fullnode.mainnet.aptoslabs.com/v1/view'
    headers = {
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'en-US,en;q=0.9,uk;q=0.8',
        'cache-control': 'no-cache',
        'content-type': 'application/json'
    }
    data = json.dumps({
        "arguments":[f"{wallet}"],
        "function":"0x53a30a6e5936c0a4c5140daed34de39d17ca7fcae08f947c02e979cef98a3719::claim::get_user_stats","type_arguments":[]})
    response = request(method='post', url=url, wallet=wallet, proxy=proxy, headers=headers, data=data)
    amount1 = response[0]
    amount2 = response[1]
    amount3 = response[2]
    add_data(wallet, amount1, amount2, amount3)
       

if __name__ == '__main__':
    for wallet in WALLETS:
        start_check(wallet)
    logger.info('Verified all data')